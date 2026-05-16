import streamlit as st
import pandas as pd
from datetime import datetime, date, time, timedelta
import io
import sqlite3
from pathlib import Path

from reportlab.lib.pagesizes import letter
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, Image as RLImage)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

# ── Base de datos SQLite ─────────────────────────────────────────────────────
DB_PATH = Path(__file__).parent / "data" / "pptmal03.db"
DB_PATH.parent.mkdir(exist_ok=True)

def get_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS registros (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                hora         TEXT NOT NULL,
                fecha        TEXT NOT NULL,
                turno        TEXT NOT NULL DEFAULT 'Turno 1',
                temp_motor   REAL NOT NULL DEFAULT 0,
                temp_reductor REAL NOT NULL DEFAULT 0,
                estado       TEXT NOT NULL
            )
        """)
        for col, definition in [
            ("turno",         "TEXT NOT NULL DEFAULT 'Turno 1'"),
            ("operario",      "TEXT NOT NULL DEFAULT ''"),
            ("temp_motor",    "REAL NOT NULL DEFAULT 0"),
            ("temp_reductor", "REAL NOT NULL DEFAULT 0"),
        ]:
            try:
                conn.execute(f"ALTER TABLE registros ADD COLUMN {col} {definition}")
            except Exception:
                pass
        # Migración: si existe columna "temp" (nombre anterior), copiar a temp_motor
        try:
            conn.execute("UPDATE registros SET temp_motor = temp WHERE temp_motor = 0 AND temp IS NOT NULL")
        except Exception:
            pass

def cargar_registros() -> pd.DataFrame:
    with get_conn() as conn:
        df = pd.read_sql("SELECT * FROM registros ORDER BY id", conn)
    if not df.empty:
        df['datetime'] = pd.to_datetime(df['fecha'] + ' ' + df['hora'])
    return df

def insertar_registro(hora, fecha, turno, operario, temp_motor, temp_reductor, estado):
    with get_conn() as conn:
        conn.execute(
            """INSERT INTO registros
               (hora, fecha, turno, operario, temp_motor, temp_reductor, estado)
               VALUES (?,?,?,?,?,?,?)""",
            (hora, fecha, turno, operario, temp_motor, temp_reductor, estado)
        )

def actualizar_registro(row_id, temp_motor, temp_reductor, estado):
    with get_conn() as conn:
        conn.execute(
            "UPDATE registros SET temp_motor=?, temp_reductor=?, estado=? WHERE id=?",
            (temp_motor, temp_reductor, estado, row_id)
        )

def eliminar_registros(ids: list):
    with get_conn() as conn:
        conn.executemany("DELETE FROM registros WHERE id=?", [(i,) for i in ids])

init_db()

# ── Lógica de turnos ─────────────────────────────────────────────────────────
TURNOS = {
    'Turno 1  (06:00 – 13:59)': (time(6,  0), time(13, 59, 59)),
    'Turno 2  (14:00 – 21:59)': (time(14, 0), time(21, 59, 59)),
    'Turno 3  (22:00 – 05:59)': None,
}
TURNO_KEYS = list(TURNOS.keys())
TURNO_HORA = {
    'Turno 1  (06:00 – 13:59)': '06:00:00',
    'Turno 2  (14:00 – 21:59)': '14:00:00',
    'Turno 3  (22:00 – 05:59)': '22:00:00',
}

def rango_turno(turno_key: str, fecha_desde: date, fecha_hasta: date):
    if turno_key.startswith('Turno 1'):
        h_ini, h_fin = TURNOS[turno_key]
        return datetime.combine(fecha_desde, h_ini), datetime.combine(fecha_hasta, h_fin)
    elif turno_key.startswith('Turno 2'):
        h_ini, h_fin = TURNOS[turno_key]
        return datetime.combine(fecha_desde, h_ini), datetime.combine(fecha_hasta, h_fin)
    else:
        return (datetime.combine(fecha_desde - timedelta(days=1), time(22, 0)),
                datetime.combine(fecha_hasta, time(5, 59, 59)))

def desc_rango_turno(turno_key, fecha_desde, fecha_hasta):
    label = turno_key.split('(')[0].strip()
    if turno_key.startswith('Turno 3'):
        return (f"{label} · "
                f"{(fecha_desde - timedelta(days=1)).strftime('%d/%m/%Y')} 22:00"
                f" → {fecha_hasta.strftime('%d/%m/%Y')} 05:59")
    h_ini = turno_key.split('(')[1].split('–')[0].strip()
    h_fin = turno_key.split('–')[1].replace(')', '').strip()
    return (f"{label} · "
            f"{fecha_desde.strftime('%d/%m/%Y')} {h_ini}"
            f" → {fecha_hasta.strftime('%d/%m/%Y')} {h_fin}")

# ── Encabezado ───────────────────────────────────────────────────────────────
col1, col2 = st.columns([3, 1], vertical_alignment='center', gap="small")
with col1:
    st.title("MESA DE ALIMENTACION")
with col2:
    st.subheader('PPTMAL03')

# ── Nuevo registro ────────────────────────────────────────────────────────────
st.subheader('Nuevo registro')

OPERARIOS = [
    'Seleccionar operario...',
    'Bladimir Recalde',
    'Adriano Paz',
    'Faber Rios',
    'Javier Henao',
    'Julian Palomeque',
    'Julian Reina',
    'Wendy Perez',
    'Robert Delgado',
    'Cristhian Caicedo',
    'Brayan Franco',
    
]

opciones_estado = ['Tensionada', 'Destensionada', 'Rota']
df_todos = cargar_registros()
ultimo_estado = df_todos['estado'].iloc[-1] if not df_todos.empty else 'Tensionada'

r1c1, r1c2, r1c3 = st.columns(3, gap="small")
with r1c1:
    fecha_reg = st.date_input('Fecha del registro', value=date.today(), key='fecha_reg')
with r1c2:
    turno_reg = st.selectbox('Turno', options=TURNO_KEYS, key='turno_reg')
with r1c3:
    operario_reg = st.selectbox('Operario', options=OPERARIOS, key='operario_reg')

r2c1, r2c2, r2c3 = st.columns(3, gap="small")
with r2c1:
    temp_motor = st.number_input(
        'Temperatura Motor (°C)', min_value=0, max_value=150, step=1, key='temp_motor')
with r2c2:
    temp_reductor = st.number_input(
        'Temperatura Reductor (°C)', min_value=0, max_value=150, step=1, key='temp_reductor')
with r2c3:
    estado = st.selectbox(
        'Cadena de Alimentación',
        options=opciones_estado,
        index=opciones_estado.index(ultimo_estado) if ultimo_estado in opciones_estado else 0,
        key='estado_reg'
    )

if st.button('💾 Guardar registro', use_container_width=True):
    if operario_reg == OPERARIOS[0]:
        st.warning('⚠️ Selecciona un operario antes de guardar.')
    else:
        turno_label = turno_reg.split('(')[0].strip()
        insertar_registro(
            hora=TURNO_HORA[turno_reg],
            fecha=fecha_reg.strftime('%Y-%m-%d'),
            turno=turno_label,
            operario=operario_reg,
            temp_motor=temp_motor,
            temp_reductor=temp_reductor,
            estado=estado
        )
        st.success(f'Guardado: {fecha_reg.strftime("%d/%m/%Y")} · {turno_label} · {operario_reg} ✓')
        st.rerun()

# ── Recargar ──────────────────────────────────────────────────────────────────
df_todos = cargar_registros()
if df_todos.empty:
    st.info("Aún no hay registros. Guarda el primero.")
    st.stop()

# ── Filtro ───────────────────────────────────────────────────────────────────
st.divider()
st.subheader('🔍 Visualizar datos por fecha y turno')

fecha_min = df_todos['datetime'].min().date()
fecha_max = df_todos['datetime'].max().date()

fc1, fc2 = st.columns(2)
with fc1:
    fecha_desde = st.date_input('Fecha desde', value=fecha_min,
                                min_value=fecha_min, max_value=fecha_max, key='fecha_desde')
with fc2:
    fecha_hasta = st.date_input('Fecha hasta', value=fecha_max,
                                min_value=fecha_min, max_value=fecha_max, key='fecha_hasta')

OPCION_TODOS = 'Todos los turnos'
turno_sel = st.radio(
    'Turno a visualizar',
    options=[OPCION_TODOS] + TURNO_KEYS,
    horizontal=True,
    key='turno_sel'
)

if fecha_desde > fecha_hasta:
    st.warning('⚠️ La fecha de inicio es posterior a la de fin.')
    st.stop()

if turno_sel == OPCION_TODOS:
    dt_inicio = datetime.combine(fecha_desde, time(0, 0, 0))
    dt_fin    = datetime.combine(fecha_hasta, time(23, 59, 59))
    desc_rango = (f"Todos los turnos · "
                  f"{fecha_desde.strftime('%d/%m/%Y')} → "
                  f"{fecha_hasta.strftime('%d/%m/%Y')}")
else:
    dt_inicio, dt_fin = rango_turno(turno_sel, fecha_desde, fecha_hasta)
    desc_rango = desc_rango_turno(turno_sel, fecha_desde, fecha_hasta)

df = df_todos[
    (df_todos['datetime'] >= dt_inicio) &
    (df_todos['datetime'] <= dt_fin)
].copy()

st.caption(f"🕐 Rango efectivo: **{desc_rango}**")

# ── Métricas ──────────────────────────────────────────────────────────────────
col_m1, col_m2, col_m3, col_m4 = st.columns(4)
col_m1.metric('Total registros', len(df_todos))
col_m2.metric('Registros en rango', len(df))
col_m3.metric('Prom. Motor',
              f"{df['temp_motor'].mean():.1f} °C" if not df.empty else '—')
col_m4.metric('Prom. Reductor',
              f"{df['temp_reductor'].mean():.1f} °C" if not df.empty else '—')

if df.empty:
    st.info('No hay registros en el rango seleccionado.')
    st.stop()

# ── Gráfica ───────────────────────────────────────────────────────────────────
st.divider()
st.subheader('Historial de Temperaturas')

df_chart = df.copy()
df_chart['Etiqueta'] = df_chart['fecha'] + ' · ' + df_chart['turno']
df_plot = df_chart.set_index('Etiqueta')[['temp_motor', 'temp_reductor']]
df_plot.columns = ['Motor (°C)', 'Reductor (°C)']
st.line_chart(df_plot)

# ── Tabla editable ────────────────────────────────────────────────────────────
st.divider()
st.subheader('Registros')
st.caption("Edita Temp. Motor, Temp. Reductor y Estado directamente. "
           "Marca ☑ para seleccionar filas a eliminar.")

df_edit = df[['id', 'fecha', 'turno', 'operario', 'hora', 'temp_motor', 'temp_reductor', 'estado']].copy()
df_edit.rename(columns={
    'id': 'ID', 'fecha': 'Fecha', 'turno': 'Turno', 'operario': 'Operario', 'hora': 'Hora',
    'temp_motor': 'T° Motor', 'temp_reductor': 'T° Reductor', 'estado': 'Estado'
}, inplace=True)
df_edit.insert(0, 'Eliminar', False)

edited = st.data_editor(
    df_edit,
    column_config={
        'Eliminar':    st.column_config.CheckboxColumn('🗑 Eliminar'),
        'ID':          st.column_config.NumberColumn('ID', disabled=True),
        'Fecha':       st.column_config.TextColumn('Fecha', disabled=True),
        'Turno':       st.column_config.TextColumn('Turno', disabled=True),
        'Operario':    st.column_config.TextColumn('Operario', disabled=True),
        'Hora':        st.column_config.TextColumn('Hora', disabled=True),
        'T° Motor':    st.column_config.NumberColumn('T° Motor (°C)',
                           min_value=0, max_value=150, step=1),
        'T° Reductor': st.column_config.NumberColumn('T° Reductor (°C)',
                           min_value=0, max_value=150, step=1),
        'Estado':      st.column_config.SelectboxColumn('Estado',
                           options=opciones_estado),
    },
    use_container_width=True,
    hide_index=True,
    key='tabla_editor'
)

col_guardar, col_eliminar = st.columns(2)

with col_guardar:
    if st.button('💾 Guardar cambios', use_container_width=True):
        cambios = 0
        for _, row in edited.iterrows():
            orig = df[df['id'] == row['ID']].iloc[0]
            if (row['T° Motor']    != orig['temp_motor'] or
                    row['T° Reductor'] != orig['temp_reductor'] or
                    row['Estado']      != orig['estado']):
                actualizar_registro(row['ID'], row['T° Motor'],
                                    row['T° Reductor'], row['Estado'])
                cambios += 1
        if cambios:
            st.success(f'{cambios} registro(s) actualizado(s) ✓')
            st.rerun()
        else:
            st.info('Sin cambios detectados.')

with col_eliminar:
    filas_a_eliminar = edited[edited['Eliminar'] == True]['ID'].tolist()
    if st.button(
        f'🗑 Eliminar seleccionados ({len(filas_a_eliminar)})',
        use_container_width=True, type='primary',
        disabled=len(filas_a_eliminar) == 0
    ):
        eliminar_registros(filas_a_eliminar)
        st.success(f'{len(filas_a_eliminar)} registro(s) eliminado(s) ✓')
        st.rerun()

# ── Generadores de archivos ───────────────────────────────────────────────────
def make_chart_image(df: pd.DataFrame) -> io.BytesIO:
    etiquetas = (df['fecha'] + ' · ' + df['turno']).tolist()
    fig, ax = plt.subplots(figsize=(8, 3.5))
    ax.plot(etiquetas, df['temp_motor'],    marker='o', color='#1f77b4',
            linewidth=2, label='Motor')
    ax.plot(etiquetas, df['temp_reductor'], marker='s', color='#ff7f0e',
            linewidth=2, label='Reductor')
    ax.set_title('Temperaturas – Mesa de Alimentación', fontsize=11)
    ax.set_xlabel('Fecha · Turno')
    ax.set_ylabel('°C')
    ax.legend()
    ax.grid(True, linestyle='--', alpha=0.5)
    plt.xticks(rotation=30, ha='right', fontsize=7)
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150)
    plt.close(fig)
    buf.seek(0)
    return buf

def generar_pdf(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('titulo', parent=styles['Title'],
                                  fontSize=16, spaceAfter=6)
    sub_style = ParagraphStyle('sub', parent=styles['Heading2'],
                               fontSize=11, spaceAfter=4)
    story = []
    story.append(Paragraph("PTM: Lista de Chequeo", titulo_style))
    story.append(Paragraph("Mesa de Alimentación – PPTMAL03", sub_style))
    story.append(Paragraph(
        f"Informe generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        styles['Normal']))
    story.append(Paragraph(f"Rango efectivo: {desc_rango}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("Historial de Temperaturas", sub_style))
    story.append(RLImage(make_chart_image(df), width=16*cm, height=6*cm))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("Registros", sub_style))

    headers = ['ID', 'Fecha', 'Turno', 'Operario', 'T° Motor', 'T° Reductor', 'Estado']
    col_w   = [1*cm, 2.5*cm, 2*cm, 3*cm, 2.5*cm, 3*cm, 3.2*cm]
    table_data = [headers]
    for _, row in df.iterrows():
        table_data.append([
            str(int(row['id'])), row['fecha'], row.get('turno', ''),
            row.get('operario', ''),
            f"{row['temp_motor']} °C", f"{row['temp_reductor']} °C", row['estado']
        ])
    tbl = Table(table_data, colWidths=col_w)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',     (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
        ('TEXTCOLOR',      (0, 0), (-1, 0), colors.white),
        ('FONTNAME',       (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',       (0, 0), (-1, -1), 8),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#dce6f0')]),
        ('GRID',           (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING',     (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    return buf.read()

def generar_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='1F4E79')
    center      = Alignment(horizontal='center', vertical='center')
    thin        = Side(style='thin', color='BFBFBF')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill    = PatternFill('solid', start_color='DCE6F0')

    for merge, val, fnt in [
        ('A1:G1', 'Mesa de Alimentación – PPTMAL03',
         Font(name='Arial', bold=True, size=14, color='1F4E79')),
        ('A2:G2', f"Informe: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
         Font(name='Arial', size=10, italic=True)),
        ('A3:G3', f"Rango efectivo: {desc_rango}",
         Font(name='Arial', size=9, italic=True, color='555555')),
    ]:
        ws.merge_cells(merge)
        cell = ws[merge.split(':')[0]]
        cell.value, cell.font, cell.alignment = val, fnt, center

    headers = ['ID', 'Fecha', 'Turno', 'Operario', 'T° Motor (°C)', 'T° Reductor (°C)', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = (
            header_font, header_fill, center, border)

    for i, (_, row) in enumerate(df.iterrows(), start=6):
        fill = alt_fill if i % 2 == 0 else PatternFill()
        vals = [int(row['id']), row['fecha'], row.get('turno', ''),
                row.get('operario', ''), row['temp_motor'], row['temp_reductor'], row['estado']]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font, cell.alignment, cell.border, cell.fill = (
                Font(name='Arial', size=10), center, border, fill)

    for col_l, w in zip('ABCDEFG', [6, 14, 12, 18, 16, 16, 18]):
        ws.column_dimensions[col_l].width = w

    # Hoja gráfica con dos series
    ws_g = wb.create_sheet("Grafica")
    ws_g['A1'] = 'Fecha · Turno'
    ws_g['B1'] = 'T° Motor (°C)'
    ws_g['C1'] = 'T° Reductor (°C)'
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        ws_g.cell(row=i, column=1, value=f"{row['fecha']} · {row.get('turno', row['hora'])}")
        ws_g.cell(row=i, column=2, value=row['temp_motor'])
        ws_g.cell(row=i, column=3, value=row['temp_reductor'])

    chart = LineChart()
    chart.title        = "Temperaturas Motor y Reductor"
    chart.y_axis.title = "°C"
    chart.x_axis.title = "Fecha · Turno"
    chart.style, chart.width, chart.height = 10, 22, 12
    chart.add_data(Reference(ws_g, min_col=2, max_col=3,
                             min_row=1, max_row=len(df)+1),
                   titles_from_data=True)
    chart.set_categories(Reference(ws_g, min_col=1,
                                   min_row=2, max_row=len(df)+1))
    ws_g.add_chart(chart, "E2")

    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── Botones de descarga ───────────────────────────────────────────────────────
st.divider()
st.caption(f"Los informes incluyen solo los **{len(df)}** registros del rango seleccionado.")
col_pdf, col_xlsx = st.columns(2)
ts = datetime.now().strftime('%Y%m%d_%H%M%S')

with col_pdf:
    st.download_button(
        label="📄 Descargar Informe PDF",
        data=generar_pdf(df),
        file_name=f"informe_PPTMAL03_{ts}.pdf",
        mime="application/pdf",
        use_container_width=True
    )
with col_xlsx:
    st.download_button(
        label="📊 Descargar Informe Excel",
        data=generar_excel(df),
        file_name=f"informe_PPTMAL03_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )